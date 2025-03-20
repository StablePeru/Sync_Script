import json
import os
from typing import Dict, Any, List, Optional

from PyQt5.QtWidgets import (QWidget, QPushButton, QHBoxLayout, QVBoxLayout, 
                            QProgressBar, QTextEdit, QMessageBox, QFileDialog,
                            QTableWidget)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QTextCursor

from src.core.audio_sync import SyncWorker

class SyncPanel(QWidget):
    # Constantes
    BUTTON_START_TEXT = "Iniciar Sincronización"
    BUTTON_SAVE_TEXT = "Guardar Resultados"
    BUTTON_EXPORT_TEXT = "Exportar a Excel"
    
    EXCEL_COLUMN_HEADERS = ["ID", "IN", "OUT", "PERSONAJE", "DIÁLOGO"]
    EXCEL_SHEET_NAME = "Sincronización"
    EXCEL_HIGHLIGHT_COLOR = "FFFF00"
    
    def __init__(self, parent: QWidget) -> None:
        super().__init__()
        self.parent = parent
        self.worker: Optional[SyncWorker] = None
        self.initUI()
        
    def initUI(self) -> None:
        """Inicializa la interfaz de usuario del panel de sincronización"""
        layout = QVBoxLayout()
        
        # Configurar botones
        button_layout = self._setup_buttons()
        
        # Configurar barra de progreso
        self.progress_bar = self._create_progress_bar()
        
        # Configurar área de log
        self.log_area = self._create_log_area()
        
        # Añadir componentes al layout
        layout.addLayout(button_layout)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.log_area)
        
        self.setLayout(layout)

    def _setup_buttons(self) -> QHBoxLayout:
        """Configura los botones del panel"""
        button_layout = QHBoxLayout()
        
        # Botón de inicio
        self.start_button = QPushButton(self.BUTTON_START_TEXT)
        self.start_button.clicked.connect(self.start_sync)
        self.start_button.setEnabled(False)
        
        # Botón de guardar
        self.save_button = QPushButton(self.BUTTON_SAVE_TEXT)
        self.save_button.clicked.connect(lambda: self.save_results(automatic=False))
        self.save_button.setVisible(False)
        
        # Botón de exportar a Excel
        self.export_button = QPushButton(self.BUTTON_EXPORT_TEXT)
        self.export_button.clicked.connect(lambda: self.export_to_excel(automatic=False))
        self.export_button.setVisible(False)
        
        button_layout.addWidget(self.start_button)
        button_layout.addWidget(self.save_button)
        button_layout.addWidget(self.export_button)
        
        return button_layout
    
    def _create_progress_bar(self) -> QProgressBar:
        """Crea y configura la barra de progreso"""
        progress_bar = QProgressBar()
        progress_bar.setTextVisible(True)
        progress_bar.setFormat("%p%")
        return progress_bar
    
    def _create_log_area(self) -> QTextEdit:
        """Crea y configura el área de log"""
        log_area = QTextEdit()
        log_area.setReadOnly(True)
        return log_area

    def enable_start_button(self) -> None:
        """Habilita el botón de inicio de sincronización"""
        self.start_button.setEnabled(True)
        
    def disable_start_button(self) -> None:
        """Deshabilita el botón de inicio de sincronización"""
        self.start_button.setEnabled(False)
        
    def start_sync(self) -> None:
        """Inicia el proceso de sincronización en un hilo separado"""
        # Configurar rutas de salida
        script_path = self.parent.get_script_path()
        self._setup_output_paths(script_path)
            
        # Reiniciar UI
        self._reset_ui_for_sync()
        
        # Crear e iniciar el worker thread
        self._create_and_start_worker(script_path)
    
    def _setup_output_paths(self, script_path: str) -> None:
        """Configura las rutas de salida basadas en el script"""
        script_basename = os.path.splitext(os.path.basename(script_path))[0]
        output_json_path = f"{script_basename}.json"
        self.parent.set_output_path(output_json_path)
    
    def _reset_ui_for_sync(self) -> None:
        """Reinicia la UI para comenzar una nueva sincronización"""
        self.log_area.clear()
        self.parent.results_panel.results_widget.setRowCount(0)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.start_button.setEnabled(False)
        self.save_button.setVisible(False)
        self.export_button.setVisible(False)
    
    def _create_and_start_worker(self, script_path: str) -> None:
        """Crea e inicia el worker thread para la sincronización"""
        self.worker = SyncWorker(
            self.parent.get_audio_path(), 
            script_path
        )
        self.worker.progress_update.connect(self.update_log)
        self.worker.progress_percent.connect(self.update_progress)
        self.worker.finished_signal.connect(self.sync_finished)
        self.worker.error_signal.connect(self.sync_error)
        self.worker.start()
    
    def update_progress(self, percent: int) -> None:
        """Actualiza la barra de progreso con el porcentaje recibido"""
        self.progress_bar.setValue(percent)
        self.progress_bar.setFormat(f"{percent}%")
        self.progress_bar.setTextVisible(True)
        
    def update_log(self, message: str) -> None:
        """Actualiza el área de registro con nuevos mensajes"""
        self.log_area.append(message)
        # Auto-scroll to bottom
        cursor = self.log_area.textCursor()
        cursor.movePosition(QTextCursor.End)
        self.log_area.setTextCursor(cursor)
        
    def sync_finished(self, json_data: Dict[str, Any]) -> None:
        """Maneja la finalización exitosa del proceso de sincronización"""
        self.progress_bar.setValue(100)
        
        # Procesar y mostrar resultados
        self._process_sync_results(json_data)
        
        # Actualizar UI
        self._update_ui_after_sync()
    
    def _process_sync_results(self, json_data: Dict[str, Any]) -> None:
        """Procesa y muestra los resultados de la sincronización"""
        # Guardar datos para uso posterior
        self.parent.set_sync_results(json_data)
        
        # Mostrar resultados en la tabla
        self.parent.display_results(json_data)
        
        # Guardar y exportar automáticamente
        self.save_results(automatic=True)
        self.export_to_excel(automatic=True)
    
    def _update_ui_after_sync(self) -> None:
        """Actualiza la UI después de completar la sincronización"""
        self.update_log("\n¡Sincronización completada!")
        self.update_log("Los resultados se han guardado automáticamente.")
        self.update_log("Revise los resultados en la pestaña 'Resultados de Sincronización'.")
        
        # Cambiar a la pestaña de resultados
        self.parent.switch_to_results_tab()
        
        # Actualizar estado de los botones
        self.save_button.setVisible(True)
        self.export_button.setVisible(True)
        self.start_button.setEnabled(True)
    
    def sync_error(self, error_message: str) -> None:
        """Maneja los errores durante el proceso de sincronización"""
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.update_log(f"ERROR: {error_message}")
        self.start_button.setEnabled(True)
        self.save_button.setVisible(False)
        self.export_button.setVisible(False)
        
        QMessageBox.critical(self, "Error", 
                            f"Ha ocurrido un error durante la sincronización:\n{error_message}")
    
    def save_results(self, automatic: bool = False) -> None:
        """
        Guarda los resultados de sincronización en un archivo JSON
        
        Args:
            automatic: Si es True, guarda automáticamente sin mostrar diálogos
        """
        try:
            self._save_json_file(automatic)
        except Exception as e:
            self._handle_save_error(str(e), automatic)
    
    def _save_json_file(self, automatic: bool) -> None:
        """Guarda los resultados en un archivo JSON"""
        output_path = self.parent.get_output_path()
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.parent.get_sync_results(), f, indent=4, ensure_ascii=False)
            
        self.update_log(f"Archivo JSON guardado en: {output_path}")
        
        if not automatic:
            QMessageBox.information(self, "Guardado Completado", 
                                  f"Los resultados se han guardado correctamente en:\n{output_path}")
            self.save_button.setVisible(False)
    
    def _handle_save_error(self, error_msg: str, automatic: bool) -> None:
        """Maneja errores durante el guardado de archivos"""
        error_msg = f"No se pudo guardar el archivo JSON: {error_msg}"
        self.update_log(f"ERROR: {error_msg}")
        if not automatic:
            QMessageBox.critical(self, "Error al Guardar", error_msg)
    
    def export_to_excel(self, automatic: bool = False) -> None:
        """
        Exporta los resultados a un archivo Excel (XLSX) con el mismo formato y colores
        
        Args:
            automatic: Si es True, exporta automáticamente sin mostrar diálogos
        """
        try:
            self._check_excel_dependencies()
            data = self._get_table_data()
            file_path = self._get_excel_file_path(automatic)
            
            if file_path:
                self._create_excel_file(data, file_path, automatic)
                
        except ImportError as e:
            self._handle_excel_import_error(automatic)
        except Exception as e:
            self._handle_excel_export_error(str(e), automatic)
    
    def _check_excel_dependencies(self) -> None:
        """Verifica que las dependencias para Excel estén instaladas"""
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
    
    def _get_table_data(self) -> List[List[str]]:
        """Obtiene los datos de la tabla de resultados"""
        data = []
        results_widget = self.parent.results_panel.results_widget
        for row in range(results_widget.rowCount()):
            row_data = []
            for col in range(results_widget.columnCount()):
                item = results_widget.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)
        return data
    
    def _get_excel_file_path(self, automatic: bool) -> Optional[str]:
        """Determina la ruta del archivo Excel"""
        if automatic:
            # Usar la ruta del JSON pero cambiando la extensión
            json_path = self.parent.get_output_path()
            file_path = json_path.replace('.json', '.xlsx')
        else:
            # Solicitar ubicación para guardar
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Guardar como Excel", "", 
                "Archivos Excel (*.xlsx)"
            )
        
        if file_path and not file_path.endswith('.xlsx'):
            file_path += '.xlsx'
            
        return file_path
    
    def _create_excel_file(self, data: List[List[str]], file_path: str, automatic: bool) -> None:
        """Crea y guarda el archivo Excel con los datos proporcionados"""
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Crear DataFrame
        df = pd.DataFrame(data, columns=self.EXCEL_COLUMN_HEADERS)
        
        # Crear un libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = self.EXCEL_SHEET_NAME
        
        # Añadir los datos al libro
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Aplicar formato (resaltado amarillo)
        self._apply_excel_highlighting(ws)
        
        # Ajustar el ancho de las columnas
        self._adjust_excel_column_widths(ws)
        
        # Guardar el archivo
        wb.save(file_path)
        self.update_log(f"Archivo Excel guardado en: {file_path}")
        
        if not automatic:
            QMessageBox.information(self, "Exportación Completada", 
                                  f"Los resultados se han exportado correctamente a:\n{file_path}")
    
    def _apply_excel_highlighting(self, worksheet) -> None:
        """Aplica el resaltado a las filas en Excel"""
        from openpyxl.styles import PatternFill
        
        yellow_fill = PatternFill(
            start_color=self.EXCEL_HIGHLIGHT_COLOR, 
            end_color=self.EXCEL_HIGHLIGHT_COLOR, 
            fill_type="solid"
        )
        
        results_widget = self.parent.results_panel.results_widget
        for row in range(results_widget.rowCount()):
            # Verificar si la primera celda de la fila está resaltada
            first_item = results_widget.item(row, 0)
            if first_item and first_item.background().color().name() == "#ffff00":
                # Aplicar resaltado a toda la fila en Excel
                for col in range(1, len(self.EXCEL_COLUMN_HEADERS) + 1):
                    worksheet.cell(row=row+2, column=col).fill = yellow_fill  # +2 porque la fila 1 es el encabezado
    
    def _adjust_excel_column_widths(self, worksheet) -> None:
        """Ajusta el ancho de las columnas en Excel"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _handle_excel_import_error(self, automatic: bool) -> None:
        """Maneja errores de importación de módulos para Excel"""
        error_msg = "Para exportar a Excel, necesita instalar pandas y openpyxl:\npip install pandas openpyxl"
        self.update_log(f"ERROR: {error_msg}")
        if not automatic:
            QMessageBox.warning(self, "Módulos Faltantes", error_msg)
    
    def _handle_excel_export_error(self, error_msg: str, automatic: bool) -> None:
        """Maneja errores durante la exportación a Excel"""
        error_msg = f"No se pudo exportar a Excel: {error_msg}"
        self.update_log(f"ERROR: {error_msg}")
        if not automatic:
            QMessageBox.critical(self, "Error al Exportar", error_msg)