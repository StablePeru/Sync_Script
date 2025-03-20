import os
import json
from PyQt5.QtWidgets import (QMainWindow, QPushButton, QVBoxLayout, 
                            QHBoxLayout, QFileDialog, QLabel, QWidget, QProgressBar,
                            QTextEdit, QMessageBox, QSplitter, QTabWidget, QTableWidget,
                            QTableWidgetItem, QHeaderView)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor

from src.core.audio_sync import SyncWorker

class SyncApp(QMainWindow):
    """
    Ventana principal de la aplicación de sincronización
    """
    def __init__(self):
        super().__init__()
        self.audio_path = ""
        self.script_path = ""
        self.output_path = ""
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('Sincronizador de Audio y Guion')
        self.setGeometry(100, 100, 1000, 700)  # Ventana más grande para acomodar la vista previa
        
        # Main widget and layout
        main_widget = QWidget()
        main_layout = QVBoxLayout()
        
        # Audio file selection
        audio_layout = QHBoxLayout()
        self.audio_label = QLabel("Archivo de audio: No seleccionado")
        audio_button = QPushButton("Seleccionar Audio")
        audio_button.clicked.connect(self.select_audio)
        audio_layout.addWidget(self.audio_label)
        audio_layout.addWidget(audio_button)
        
        # Script file selection
        script_layout = QHBoxLayout()
        self.script_label = QLabel("Archivo de guion: No seleccionado")
        script_button = QPushButton("Seleccionar Guion")
        script_button.clicked.connect(self.select_script)
        script_layout.addWidget(self.script_label)
        script_layout.addWidget(script_button)
        
        # Output file selection
        output_layout = QHBoxLayout()
        self.output_label = QLabel("Archivo de salida: output.json")
        output_button = QPushButton("Seleccionar Salida")
        output_button.clicked.connect(self.select_output)
        output_layout.addWidget(self.output_label)
        output_layout.addWidget(output_button)
        
        # Start button
        self.start_button = QPushButton("Iniciar Sincronización")
        self.start_button.clicked.connect(self.start_sync)
        self.start_button.setEnabled(False)
        
        # Save button (initially hidden)
        self.save_button = QPushButton("Guardar Resultados")
        self.save_button.clicked.connect(self.save_results)
        self.save_button.setVisible(False)
        
        # Export to Excel button (initially hidden)
        self.export_button = QPushButton("Exportar a Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setVisible(False)
        
        # Button layout
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.start_button)
        button_layout.addWidget(self.save_button)
        button_layout.addWidget(self.export_button)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(False)
        
        # Crear un splitter para dividir la interfaz
        splitter = QSplitter(Qt.Vertical)
        
        # Crear un widget de pestañas para la vista previa, resultados y log
        self.tab_widget = QTabWidget()
        
        # Vista previa del guion
        self.script_preview = QTextEdit()
        self.script_preview.setReadOnly(True)
        self.script_preview.setPlaceholderText("El contenido del guion se mostrará aquí cuando se cargue un archivo.")
        self.tab_widget.addTab(self.script_preview, "Vista Previa del Guion")
        
        # Resultados de sincronización
        self.results_widget = QTableWidget()
        self.results_widget.setColumnCount(5)
        self.results_widget.setHorizontalHeaderLabels(["ID", "IN", "OUT", "PERSONAJE", "DIÁLOGO"])
        self.results_widget.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)  # Diálogo se expande
        self.tab_widget.addTab(self.results_widget, "Resultados de Sincronización")
        
        # Log area
        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.tab_widget.addTab(self.log_area, "Log de Procesamiento")
        
        # Añadir el widget de pestañas al splitter
        splitter.addWidget(self.tab_widget)
        
        # Add all widgets to main layout
        main_layout.addLayout(audio_layout)
        main_layout.addLayout(script_layout)
        main_layout.addLayout(output_layout)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.progress_bar)
        main_layout.addWidget(splitter, 1)  # El splitter toma el espacio restante
        
        main_widget.setLayout(main_layout)
        self.setCentralWidget(main_widget)
        
    def select_audio(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo de audio", "", 
            "Archivos de audio (*.wav *.mp3 *.m4a *.ogg)"
        )
        if file_path:
            self.audio_path = file_path
            self.audio_label.setText(f"Archivo de audio: {os.path.basename(file_path)}")
            self.check_start_button()
            
    def select_script(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo de guion", "", 
            "Archivos de texto (*.txt)"
        )
        if file_path:
            self.script_path = file_path
            self.script_label.setText(f"Archivo de guion: {os.path.basename(file_path)}")
            
            # Cargar y mostrar el contenido del guion en la vista previa
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    script_content = f.read()
                self.script_preview.setText(script_content)
                self.script_preview.moveCursor(self.script_preview.textCursor().Start)
            except Exception as e:
                self.script_preview.setText(f"Error al cargar el guion: {str(e)}")
                QMessageBox.warning(self, "Error de Lectura", 
                                   f"No se pudo leer el archivo de guion:\n{str(e)}")
            
            self.check_start_button()
            
    def select_output(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Guardar archivo JSON", "output.json", 
            "Archivos JSON (*.json)"
        )
        if file_path:
            self.output_path = file_path
            self.output_label.setText(f"Archivo de salida: {os.path.basename(file_path)}")
    
    def check_start_button(self):
        """
        Habilita el botón de inicio si se han seleccionado los archivos necesarios
        """
        if self.audio_path and self.script_path:
            self.start_button.setEnabled(True)
        else:
            self.start_button.setEnabled(False)
            
    def start_sync(self):
        """
        Inicia el proceso de sincronización en un hilo separado
        """
        if not self.output_path:
            self.output_path = "output.json"
            
        self.log_area.clear()
        self.results_widget.setRowCount(0)  # Limpiar resultados anteriores
        self.progress_bar.setRange(0, 0)  # Indeterminate progress
        self.start_button.setEnabled(False)
        self.save_button.setVisible(False)
        
        # Create and start worker thread
        self.worker = SyncWorker(self.audio_path, self.script_path)
        self.worker.progress_update.connect(self.update_log)
        self.worker.finished_signal.connect(self.sync_finished)
        self.worker.error_signal.connect(self.sync_error)
        self.worker.start()
        
    def update_log(self, message):
        """
        Actualiza el área de registro con nuevos mensajes
        """
        self.log_area.append(message)
        # Auto-scroll to bottom
        cursor = self.log_area.textCursor()
        cursor.movePosition(cursor.End)
        self.log_area.setTextCursor(cursor)
        
    def sync_finished(self, json_data):
        """
        Maneja la finalización exitosa del proceso de sincronización
        """
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(100)
        
        # Guardar los datos para uso posterior
        self.sync_results = json_data
        
        # Mostrar los resultados en la tabla
        self.display_results(json_data)
        
        self.update_log(f"\n¡Sincronización completada!")
        self.update_log(f"Revise los resultados en la pestaña 'Resultados de Sincronización' y haga clic en 'Guardar Resultados' para guardar el archivo.")
        
        # Cambiar a la pestaña de resultados
        self.tab_widget.setCurrentIndex(1)
        
        # Mostrar los botones de guardar y exportar, y habilitar el botón de inicio
        self.save_button.setVisible(True)
        self.export_button.setVisible(True)
        self.start_button.setEnabled(True)
        
    def display_results(self, json_data):
        """
        Muestra los resultados de sincronización en la tabla
        """
        data = json_data.get("data", [])
        self.results_widget.setRowCount(len(data))
        
        prev_in_time = None  # Para rastrear el tiempo IN anterior
        
        for row, item in enumerate(data):
            # ID
            id_item = QTableWidgetItem(str(item.get("ID", "")))
            self.results_widget.setItem(row, 0, id_item)
            
            # IN
            in_time = item.get("IN", "00:00:00:00")
            in_item = QTableWidgetItem(in_time)
            self.results_widget.setItem(row, 1, in_item)
            
            # OUT
            out_time = item.get("OUT", "00:00:00:00")
            out_item = QTableWidgetItem(out_time)
            self.results_widget.setItem(row, 2, out_item)
            
            # PERSONAJE
            character_item = QTableWidgetItem(item.get("PERSONAJE", ""))
            self.results_widget.setItem(row, 3, character_item)
            
            # DIÁLOGO
            dialogue_item = QTableWidgetItem(item.get("DIÁLOGO", ""))
            self.results_widget.setItem(row, 4, dialogue_item)
            
            # Resaltar filas sin tiempo asignado o con tiempo anterior al previo
            highlight_row = False
            if in_time == "00:00:00:00" or out_time == "00:00:00:00":
                highlight_row = True
            elif prev_in_time and self._compare_timecodes(in_time, prev_in_time) < 0:
                # Si el tiempo actual es menor que el anterior
                highlight_row = True
                
            # Aplicar resaltado a toda la fila si es necesario
            if highlight_row:
                for col in range(5):  # Resaltar todas las columnas de la fila
                    self.results_widget.item(row, col).setBackground(Qt.yellow)
            
            prev_in_time = in_time  # Actualizar el tiempo anterior
    
    def _compare_timecodes(self, tc1, tc2):
        """
        Compara dos códigos de tiempo en formato "HH:MM:SS:FF"
        Retorna: -1 si tc1 < tc2, 0 si son iguales, 1 si tc1 > tc2
        """
        # Convertir a segundos para comparación
        def tc_to_seconds(tc):
            parts = tc.split(':')
            if len(parts) != 4:
                return 0
            h, m, s, f = map(int, parts)
            return h * 3600 + m * 60 + s + f/30  # Asumiendo 30fps
            
        tc1_seconds = tc_to_seconds(tc1)
        tc2_seconds = tc_to_seconds(tc2)
        
        if tc1_seconds < tc2_seconds:
            return -1
        elif tc1_seconds > tc2_seconds:
            return 1
        else:
            return 0
            
    def export_to_excel(self):
        """
        Exporta los resultados a un archivo Excel (XLSX) con el mismo formato y colores
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Crear un DataFrame con los datos de la tabla
            data = []
            for row in range(self.results_widget.rowCount()):
                row_data = []
                for col in range(self.results_widget.columnCount()):
                    item = self.results_widget.item(row, col)
                    if item:
                        row_data.append(item.text())
                    else:
                        row_data.append("")
                data.append(row_data)
            
            # Crear DataFrame
            df = pd.DataFrame(data, columns=["ID", "IN", "OUT", "PERSONAJE", "DIÁLOGO"])
            
            # Crear un libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Sincronización"
            
            # Añadir los datos al libro
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Aplicar formato (resaltado amarillo)
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            for row in range(self.results_widget.rowCount()):
                # Verificar si la primera celda de la fila está resaltada
                first_item = self.results_widget.item(row, 0)
                if first_item and first_item.background().color().name() == "#ffff00":
                    # Aplicar resaltado a toda la fila en Excel
                    for col in range(1, 6):  # Columnas 1-5 (A-E)
                        ws.cell(row=row+2, column=col).fill = yellow_fill  # +2 porque la fila 1 es el encabezado
            
            # Ajustar el ancho de las columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Solicitar ubicación para guardar
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Guardar como Excel", "", 
                "Archivos Excel (*.xlsx)"
            )
            
            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                wb.save(file_path)
                self.update_log(f"Archivo Excel guardado en: {file_path}")
                QMessageBox.information(self, "Exportación Completada", 
                                      f"Los resultados se han exportado correctamente a:\n{file_path}")
        except ImportError:
            QMessageBox.warning(self, "Módulos Faltantes", 
                              "Para exportar a Excel, necesita instalar pandas y openpyxl:\n"
                              "pip install pandas openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Error al Exportar", 
                               f"No se pudo exportar a Excel:\n{str(e)}")
    
    def save_results(self):
        """
        Guarda los resultados de sincronización en un archivo JSON
        """
        try:
            with open(self.output_path, 'w', encoding='utf-8') as f:
                json.dump(self.sync_results, f, indent=4, ensure_ascii=False)
                
            self.update_log(f"Archivo guardado en: {self.output_path}")
            
            QMessageBox.information(self, "Guardado Completado", 
                                   f"Los resultados se han guardado correctamente en:\n{self.output_path}")
            
            # Ocultar el botón de guardar después de guardar
            self.save_button.setVisible(False)
            
        except Exception as e:
            QMessageBox.critical(self, "Error al Guardar", 
                                f"No se pudo guardar el archivo:\n{str(e)}")
    
    def sync_error(self, error_message):
        """
        Maneja los errores durante el proceso de sincronización
        """
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.update_log(f"ERROR: {error_message}")
        self.start_button.setEnabled(True)
        self.save_button.setVisible(False)
        self.export_button.setVisible(False)
        
        QMessageBox.critical(self, "Error", 
                            f"Ha ocurrido un error durante la sincronización:\n{error_message}")
